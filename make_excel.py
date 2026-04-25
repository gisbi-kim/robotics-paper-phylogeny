"""
Generate the final Excel deliverable: robotics_taxonomy.xlsx
3 sheets:
  1. Papers — all 7,477 papers with taxonomy labels
  2. Taxonomy_Tree — the taxonomy hierarchy with paper counts
  3. Stats — summary statistics
"""
import json
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/Users/gsk/Library/CloudStorage/OneDrive-개인/#. Claude/tro-계통도/robotics_taxonomy.xlsx'
INP = '/tmp/classified.json'

with open(INP, 'r', encoding='utf-8') as f:
    papers = json.load(f)

print(f"Total papers: {len(papers)}")

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# ============================================================
# Sheet 1: Papers
# ============================================================
ws1 = wb.create_sheet('Papers')
header = ['idx', 'venue', 'year', 'title', 'authors', 'citations',
          'Phylum', 'Class', 'Order', 'Genus']
ws1.append(header)

# Header styling
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092',
                          fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
for col_idx, _ in enumerate(header, start=1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for p in papers:
    ws1.append([
        p['idx'], p['venue'], p['year'], p['title'], p['authors'],
        p['citations'], p['phylum'], p['class'], p['order'], p['genus']
    ])

# Column widths
widths = [6, 12, 6, 70, 35, 8, 30, 30, 35, 30]
for i, w in enumerate(widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws1.freeze_panes = 'A2'

# Auto filter
ws1.auto_filter.ref = ws1.dimensions

# ============================================================
# Sheet 2: Taxonomy_Tree
# ============================================================
ws2 = wb.create_sheet('Taxonomy_Tree')
ws2.append(['Phylum', 'Class', 'Order', 'Genus', 'Paper Count', '% of Total'])

# Header style
for col_idx, _ in enumerate(range(6), start=1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Group by Phylum > Class > Order > Genus (4 levels)
tree = defaultdict(lambda: defaultdict(lambda: defaultdict(Counter)))
for p in papers:
    tree[p['phylum']][p['class']][p['order']][p['genus']] += 1

# Define Phylum order for output (matches TAXONOMY.md)
phylum_order = [
    'Perception & Sensing',
    'SLAM & Localization',
    'Planning',
    'Control',
    'Manipulation',
    'Locomotion',
    'Robot Design & Hardware',
    'Human-Robot Interaction',
    'Multi-Robot Systems',
    'Learning for Robotics',
    'Application Domains',
    'Theoretical Foundations',
    'Robot Software & Architecture',
    'Other / Editorial',
    'Other / Unclassified',
]

total = len(papers)

phy_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2',
                       fill_type='solid')
cls_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2',
                       fill_type='solid')

# Write Phylum > Class > Order > Genus rows
for phy in phylum_order:
    if phy not in tree:
        continue
    phy_total = sum(sum(sum(genera.values()) for genera in orders.values())
                    for orders in tree[phy].values())
    # Phylum-level summary row
    row = [phy, '', '', '', phy_total, f'{phy_total / total * 100:.1f}%']
    ws2.append(row)
    r = ws2.max_row
    for c in range(1, 7):
        ws2.cell(row=r, column=c).fill = phy_fill
        ws2.cell(row=r, column=c).font = Font(bold=True)

    # Sort classes by count desc
    classes_sorted = sorted(
        tree[phy].items(),
        key=lambda x: -sum(sum(g.values()) for g in x[1].values()))
    for cls, orders in classes_sorted:
        cls_total = sum(sum(g.values()) for g in orders.values())
        row = ['', cls, '', '', cls_total,
               f'{cls_total / total * 100:.2f}%']
        ws2.append(row)
        r = ws2.max_row
        for c in range(1, 7):
            ws2.cell(row=r, column=c).fill = cls_fill
            ws2.cell(row=r, column=c).font = Font(italic=True)

        # Sort orders by count desc
        orders_sorted = sorted(orders.items(),
                               key=lambda x: -sum(x[1].values()))
        for ord_, genera in orders_sorted:
            ord_total = sum(genera.values())
            ws2.append(['', '', ord_, '', ord_total,
                        f'{ord_total / total * 100:.2f}%'])

            # Sort genera by count desc
            genera_sorted = sorted(genera.items(), key=lambda x: -x[1])
            for gen, cnt in genera_sorted:
                ws2.append(['', '', '', gen, cnt,
                            f'{cnt / total * 100:.3f}%'])

# Column widths
for i, w in enumerate([32, 34, 38, 32, 12, 12], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A2'

# ============================================================
# Sheet 3: Stats
# ============================================================
ws3 = wb.create_sheet('Stats')

# 3a. Phylum distribution
ws3.append(['=== Phylum Distribution ==='])
ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
ws3.append(['Phylum', 'Count', '% of Total'])
for cell in ws3[2]:
    cell.font = header_font
    cell.fill = header_fill

phy_counts = Counter(p['phylum'] for p in papers)
for phy in phylum_order:
    if phy in phy_counts:
        v = phy_counts[phy]
        ws3.append([phy, v, f'{v/total*100:.1f}%'])

ws3.append([''])
ws3.append(['=== Top 30 Class Distribution ==='])
ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=14)
ws3.append(['Phylum > Class', 'Count', '% of Total'])
for cell in ws3[ws3.max_row]:
    cell.font = header_font
    cell.fill = header_fill

cls_counts = Counter(f"{p['phylum']} > {p['class']}" for p in papers)
for k, v in sorted(cls_counts.items(), key=lambda x: -x[1])[:30]:
    ws3.append([k, v, f'{v/total*100:.2f}%'])

# 3b. Year x Phylum trend
ws3.append([''])
ws3.append(['=== Year × Phylum Trend ==='])
ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=14)

years = sorted(set(p['year'] for p in papers))
ws3.append(['Year'] + phylum_order)
for cell in ws3[ws3.max_row]:
    cell.font = header_font
    cell.fill = header_fill

# Pivot: count per year per phylum
year_phy_counts = defaultdict(Counter)
for p in papers:
    year_phy_counts[p['year']][p['phylum']] += 1

for y in years:
    row = [y] + [year_phy_counts[y].get(phy, 0) for phy in phylum_order]
    ws3.append(row)

# Column widths
ws3.column_dimensions['A'].width = 50
for i in range(2, 20):
    ws3.column_dimensions[get_column_letter(i)].width = 12

ws3.freeze_panes = 'A2'

# ============================================================
# Save
# ============================================================
wb.save(OUT)
print(f"\nSaved: {OUT}")

# Final summary print
print("\n=== FINAL Phylum Distribution ===")
for phy in phylum_order:
    if phy in phy_counts:
        v = phy_counts[phy]
        print(f"  {v:>5}  {phy} ({v/total*100:.1f}%)")
